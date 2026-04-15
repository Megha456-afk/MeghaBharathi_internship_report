import requests
import pandas as pd
import numpy as np
import os
import smtplib
import warnings
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import LabelEncoder, StandardScaler

warnings.filterwarnings("ignore")

# =========================
# CONFIG — FIXED PATHS
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_DIR  = os.path.join(BASE_DIR, "input")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
TEXT_DIR   = os.path.join(BASE_DIR, "text")

FINAL_FILE   = os.path.join(OUTPUT_DIR, "FINAL_SORTED_03022026.xlsx")
SMS_FILE     = os.path.join(INPUT_DIR,  "smsbackup.xlsx")
GP_MASTER    = os.path.join(INPUT_DIR,  "KDAH Indore GP master file.xlsx")
CONFIG_FILE  = os.path.join(TEXT_DIR,   "config.txt")
ML_FILE      = os.path.join(OUTPUT_DIR, "ML_PREDICTIONS.xlsx")


# =====================================================
# STEP 1 — REMOVE DUPLICATES
# =====================================================
def remove_duplicates():
    print("\n🧹 Removing duplicates...")
    df = pd.read_excel(SMS_FILE)
    df = df.drop_duplicates()
    df.to_excel(SMS_FILE, index=False)
    print("✅ Duplicates removed")


# =====================================================
# STEP 2 — CONFIG FILTER + APPEND
# =====================================================
def config_filter():
    print("\n⚙ Applying checkpoint filter...")

    with open(CONFIG_FILE, "r") as f:
        checkpoint_str = f.readline().split("=")[1].strip()

    checkpoint = pd.to_datetime(checkpoint_str, format="%Y-%m-%d %a %I:%M %p", errors="coerce")

    data = pd.read_excel(SMS_FILE)
    data.columns = data.columns.str.strip().str.lower()
    data["date"] = pd.to_datetime(data["date"], format="%Y-%m-%d %a %I:%M %p", errors="coerce")
    data = data.dropna(subset=["date"])

    new_records = data[data["date"] > checkpoint].sort_values("date")
    mapped = new_records[["address", "date", "body"]].copy()
    mapped.columns = ["PH_NO", "SMS_DATE", "SMS_BODY"]
    mapped["SMS_DATE"] = mapped["SMS_DATE"].dt.strftime("%Y-%m-%d %a %I:%M %p")

    try:
        final_df = pd.read_excel(FINAL_FILE, dtype=str)
    except FileNotFoundError:
        final_df = pd.DataFrame(columns=mapped.columns)

    updated = pd.concat([final_df, mapped], ignore_index=True)
    updated.to_excel(FINAL_FILE, index=False)
    print("✅ Config merge complete")


# =====================================================
# STEP 3 — GP MASTER MERGE
# =====================================================
def gp_merge():
    print("\n🔗 Merging GP master...")

    gp_master    = pd.read_excel(GP_MASTER, dtype=str)
    final_sorted = pd.read_excel(FINAL_FILE, dtype=str)

    gp_master.columns    = gp_master.columns.str.strip()
    final_sorted.columns = final_sorted.columns.str.strip()

    gp_master["PH_NO"]    = gp_master["PH_NO"].str.strip()
    final_sorted["PH_NO"] = final_sorted["PH_NO"].str.strip()

    cols = ["REFERRING_DOCTOR", "AREA", "AREA_MANAGER",
            "SALES_MANAGER", "VENDOR_CODE", "PAN_NO",
            "VENDOR_NAME", "SPECIALTY"]

    merged = final_sorted.merge(
        gp_master[["PH_NO"] + cols], on="PH_NO", how="left", suffixes=("", "_gp")
    )
    for col in cols:
        merged[col] = merged[col].fillna(merged[col + "_gp"])
        merged.drop(columns=[col + "_gp"], inplace=True)

    merged.to_excel(FINAL_FILE, index=False)
    print("✅ GP merge complete")


# =====================================================
# STEP 4 — DATE EXTRACTION
# =====================================================
def extract_date_time():
    print("\n🕒 Extracting date/time...")

    df = pd.read_excel(FINAL_FILE, dtype=str)
    temp_dt = pd.to_datetime(df["SMS_DATE"], format="%Y-%m-%d %a %I:%M %p", errors="coerce")
    df.loc[temp_dt.notna(), "S_DATE"] = temp_dt.dt.strftime("%Y-%m-%d")
    df.loc[temp_dt.notna(), "S_TIME"] = temp_dt.dt.strftime("%H:%M")
    df.to_excel(FINAL_FILE, index=False)
    print("✅ Date/time extracted")


# =====================================================
# STEP 5 — FILL BLANKS
# =====================================================
def fill_blanks():
    print("\n🧩 Filling blanks...")

    df = pd.read_excel(FINAL_FILE, dtype=str)
    df = df.fillna("NOT_MAPPED")
    df.replace("", "NOT_MAPPED", inplace=True)

    sl_numeric = pd.to_numeric(df["SL.NO"], errors="coerce")
    last_sl = int(sl_numeric.max()) if sl_numeric.notna().any() else 0
    for i in df.index:
        if pd.isna(sl_numeric[i]):
            last_sl += 1
            df.at[i, "SL.NO"] = last_sl

    df.to_excel(FINAL_FILE, index=False)
    print("✅ Blank fill complete")


# =====================================================
# STEP 6 (NEW) — FEATURE ENGINEERING
#   Prepares structured ML features from cleaned data
# =====================================================
def feature_engineering(df):
    """
    Creates ML-ready features from the cleaned FINAL_SORTED data.
    Returns a feature DataFrame alongside the original.
    """
    print("\n🔧 Engineering features for ML...")

    df = df.copy()

    # Parse date
    df["S_DATE"] = pd.to_datetime(df["S_DATE"], errors="coerce")
    df = df.dropna(subset=["S_DATE"])

    # Time features
    df["WEEK"]       = df["S_DATE"].dt.to_period("W").apply(lambda r: r.start_time)
    df["MONTH"]      = df["S_DATE"].dt.to_period("M").apply(lambda r: r.start_time)
    df["DAY_OF_WEEK"] = df["S_DATE"].dt.dayofweek   # 0=Mon … 6=Sun
    df["HOUR"] = pd.to_numeric(
        df["S_TIME"].str.split(":").str[0], errors="coerce"
    ).fillna(0).astype(int)

    # Encode SPECIALTY as numeric (for clustering)
    le = LabelEncoder()
    df["SPECIALTY_ENC"] = le.fit_transform(
        df["SPECIALTY"].fillna("UNKNOWN").astype(str)
    )

    print("✅ Feature engineering complete")
    return df


# =====================================================
# STEP 7 (NEW) — REFERRAL TREND PREDICTION
#   Uses Linear Regression to forecast weekly referrals
#   per doctor for the NEXT 4 weeks
# =====================================================
def predict_referral_trends(df):
    """
    For each REFERRING_DOCTOR, trains a simple Linear Regression
    on weekly referral counts and predicts the next 4 weeks.

    ML concept  : Supervised learning — regression
    Algorithm   : LinearRegression (sklearn)
    Input feat. : Week index (1, 2, 3 … n)
    Target      : Referral count per week
    """
    print("\n📈 Predicting referral trends (Linear Regression)...")

    results = []

    doctors = df["REFERRING_DOCTOR"].unique()
    for doctor in doctors:
        if doctor in ("NOT_MAPPED", "", None):
            continue

        doc_df = df[df["REFERRING_DOCTOR"] == doctor].copy()
        weekly = (
            doc_df.groupby("WEEK").size()
            .reset_index(name="REFERRAL_COUNT")
            .sort_values("WEEK")
        )

        # Need at least 3 data points to fit a line
        if len(weekly) < 3:
            continue

        weekly["WEEK_IDX"] = range(len(weekly))
        X = weekly[["WEEK_IDX"]].values
        y = weekly["REFERRAL_COUNT"].values

        model = LinearRegression()
        model.fit(X, y)

        # Predict next 4 weeks
        last_idx   = weekly["WEEK_IDX"].max()
        last_week  = weekly["WEEK"].max()

        for i in range(1, 5):
            pred_idx  = last_idx + i
            pred_week = last_week + pd.Timedelta(weeks=i)
            pred_count = max(0, round(model.predict([[pred_idx]])[0], 1))

            results.append({
                "REFERRING_DOCTOR"  : doctor,
                "PREDICTED_WEEK"    : pred_week.strftime("%Y-%m-%d"),
                "PREDICTED_REFERRALS": pred_count,
                "TREND_SLOPE"       : round(model.coef_[0], 3),
                "TREND_DIRECTION"   : "INCREASING" if model.coef_[0] > 0.1
                                      else ("DECREASING" if model.coef_[0] < -0.1
                                      else "STABLE")
            })

    pred_df = pd.DataFrame(results)
    print(f"   → Generated predictions for {pred_df['REFERRING_DOCTOR'].nunique()} doctors")
    print("✅ Trend prediction complete")
    return pred_df


# =====================================================
# STEP 8 (NEW) — DOCTOR SEGMENTATION (CLUSTERING)
#   K-Means groups doctors into 3 segments:
#   High Referrer / Medium Referrer / Low Referrer
# =====================================================
def cluster_doctors(df):
    """
    Clusters doctors based on referral behaviour:
      - Total referral volume
      - Average referrals per week
      - Specialty encoding
      - Preferred hour of sending SMS

    ML concept  : Unsupervised learning — clustering
    Algorithm   : KMeans (sklearn), k=3
    """
    print("\n🔵 Clustering doctors by referral pattern (K-Means)...")

    doc_features = (
        df[df["REFERRING_DOCTOR"] != "NOT_MAPPED"]
        .groupby("REFERRING_DOCTOR")
        .agg(
            TOTAL_REFERRALS   = ("REFERRING_DOCTOR", "count"),
            AVG_WEEKLY        = ("WEEK",             lambda x: x.nunique()),
            SPECIALTY_ENC     = ("SPECIALTY_ENC",    "first"),
            AVG_HOUR          = ("HOUR",             "mean"),
        )
        .reset_index()
    )

    if len(doc_features) < 3:
        print("   ⚠ Not enough doctors to cluster (need ≥ 3). Skipping.")
        return pd.DataFrame()

    feature_cols = ["TOTAL_REFERRALS", "AVG_WEEKLY", "SPECIALTY_ENC", "AVG_HOUR"]
    scaler  = StandardScaler()
    X_scaled = scaler.fit_transform(doc_features[feature_cols])

    k = min(3, len(doc_features))
    kmeans = KMeans(n_clusters=k, random_state=42, n_init=10)
    doc_features["CLUSTER_ID"] = kmeans.fit_predict(X_scaled)

    # Label clusters by avg total referrals (high → low)
    cluster_means = (
        doc_features.groupby("CLUSTER_ID")["TOTAL_REFERRALS"]
        .mean()
        .sort_values(ascending=False)
    )
    label_map = {}
    labels = ["HIGH REFERRER", "MEDIUM REFERRER", "LOW REFERRER"]
    for rank, cluster_id in enumerate(cluster_means.index):
        label_map[cluster_id] = labels[rank] if rank < len(labels) else f"GROUP {rank+1}"

    doc_features["SEGMENT"] = doc_features["CLUSTER_ID"].map(label_map)

    print(f"   → Segmented {len(doc_features)} doctors into {k} groups")
    print(doc_features["SEGMENT"].value_counts().to_string())
    print("✅ Clustering complete")
    return doc_features[["REFERRING_DOCTOR", "TOTAL_REFERRALS",
                          "AVG_WEEKLY", "AVG_HOUR", "SEGMENT"]]


# =====================================================
# STEP 9 (NEW) — ANOMALY DETECTION
#   Isolation Forest flags doctors with unusual
#   drop or spike in referral count
# =====================================================
def detect_anomalies(df):
    """
    Uses Isolation Forest to detect anomalous referral patterns.
    An anomaly = doctor whose weekly referral count is
    statistically unusual (too high or too low).

    ML concept  : Unsupervised anomaly detection
    Algorithm   : IsolationForest (sklearn)
    """
    print("\n🚨 Detecting anomalies (Isolation Forest)...")

    weekly_doc = (
        df[df["REFERRING_DOCTOR"] != "NOT_MAPPED"]
        .groupby(["REFERRING_DOCTOR", "WEEK"])
        .size()
        .reset_index(name="WEEKLY_COUNT")
    )

    if len(weekly_doc) < 5:
        print("   ⚠ Not enough data for anomaly detection. Skipping.")
        return pd.DataFrame()

    iso = IsolationForest(contamination=0.1, random_state=42)
    weekly_doc["ANOMALY_SCORE"] = iso.fit_predict(weekly_doc[["WEEKLY_COUNT"]])

    # -1 = anomaly flagged by Isolation Forest
    anomalies = weekly_doc[weekly_doc["ANOMALY_SCORE"] == -1].copy()
    anomalies["FLAG"] = anomalies["WEEKLY_COUNT"].apply(
        lambda x: "UNUSUALLY HIGH" if x > weekly_doc["WEEKLY_COUNT"].mean() else "UNUSUALLY LOW"
    )
    anomalies["WEEK"] = anomalies["WEEK"].astype(str)

    print(f"   → Found {len(anomalies)} anomalous referral weeks")
    print("✅ Anomaly detection complete")
    return anomalies[["REFERRING_DOCTOR", "WEEK", "WEEKLY_COUNT", "FLAG"]]


# =====================================================
# STEP 10 (NEW) — SAVE ML RESULTS TO EXCEL
# =====================================================
def save_ml_results(predictions_df, clusters_df, anomalies_df):
    print("\n💾 Saving ML results to Excel...")

    with pd.ExcelWriter(ML_FILE, engine="openpyxl") as writer:
        if not predictions_df.empty:
            predictions_df.to_excel(writer, sheet_name="Referral_Predictions", index=False)
        if not clusters_df.empty:
            clusters_df.to_excel(writer, sheet_name="Doctor_Segments", index=False)
        if not anomalies_df.empty:
            anomalies_df.to_excel(writer, sheet_name="Anomaly_Flags", index=False)

    print(f"✅ ML results saved → {ML_FILE}")

from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

def add_graphs_to_excel(file_path):
    print("\n📊 Adding graphs to ML Excel...")

    wb = load_workbook(file_path)

    # =========================
    # 1️⃣ Referral Predictions → LINE CHART
    # =========================
    if "Referral_Predictions" in wb.sheetnames:
        ws = wb["Referral_Predictions"]

        if ws.max_row > 1:
            chart = LineChart()
            chart.title = "Referral Predictions Trend"

            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "H2")

    # =========================
    # 2️⃣ Doctor Segments → BAR CHART
    # =========================
    if "Doctor_Segments" in wb.sheetnames:
        ws = wb["Doctor_Segments"]

        if ws.max_row > 1:
            chart = BarChart()
            chart.title = "Doctor Referral Segments"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "H2")

    # =========================
    # 3️⃣ Anomaly Flags → COLUMN CHART
    # =========================
    if "Anomaly_Flags" in wb.sheetnames:
        ws = wb["Anomaly_Flags"]

        if ws.max_row > 1:
            chart = BarChart()
            chart.title = "Anomaly Detection"

            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "H2")

    wb.save(file_path)
    print("✅ Graphs added successfully!")
from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, Reference

def create_dashboard(file_path):
    print("\n📊 Creating Dashboard...")

    wb = load_workbook(file_path)

    # Create dashboard sheet
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]

    dashboard = wb.create_sheet("Dashboard")

    # =========================
    # 📈 Prediction Chart
    # =========================
    if "Referral_Predictions" in wb.sheetnames:
        ws = wb["Referral_Predictions"]

        if ws.max_row > 1:
            chart1 = LineChart()
            chart1.title = "Referral Prediction Trend"

            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(cats)

            dashboard.add_chart(chart1, "A1")

    # =========================
    # 📊 Clustering Chart
    # =========================
    if "Doctor_Segments" in wb.sheetnames:
        ws = wb["Doctor_Segments"]

        if ws.max_row > 1:
            chart2 = BarChart()
            chart2.title = "Doctor Segments"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart2.add_data(data, titles_from_data=True)
            chart2.set_categories(cats)

            dashboard.add_chart(chart2, "A15")

    # =========================
    # 🚨 Anomaly Chart
    # =========================
    if "Anomaly_Flags" in wb.sheetnames:
        ws = wb["Anomaly_Flags"]

        if ws.max_row > 1:
            chart3 = BarChart()
            chart3.title = "Anomaly Detection"

            data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)

            chart3.add_data(data, titles_from_data=True)
            chart3.set_categories(cats)

            dashboard.add_chart(chart3, "H1")
    wb.active = wb["Dashboard"]
    wb.save(file_path)
    print("✅ Dashboard created!")
# =====================================================
# STEP 11 — EMAIL (with ML report attached)
# =====================================================
def send_email():
    print("\n📧 Sending email with ML report...")

    from_email = "meghabharathi109@gmail.com"
    password   = "dcolhjmarreofylf"
    to_email   = from_email

    msg = MIMEMultipart()
    msg["From"]    = from_email
    msg["To"]      = to_email
    msg["Subject"] = "GP-SMS AI Pipeline — Predictions & Anomaly Report"

    body = """\
GP-SMS AI/ML Pipeline completed successfully.

Attached: ML_PREDICTIONS.xlsx

Sheets included:
  1. Referral_Predictions  — Next 4-week forecast per doctor (Linear Regression)
  2. Doctor_Segments       — Doctor grouping: High / Medium / Low Referrer (K-Means)
  3. Anomaly_Flags         — Unusual referral activity flagged (Isolation Forest)
"""
    msg.attach(MIMEText(body, "plain"))

    # Attach ML Excel file
    if os.path.exists(ML_FILE):
        with open(ML_FILE, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename=ML_PREDICTIONS.xlsx")
        msg.attach(part)

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(from_email, password)
    server.send_message(msg)
    server.quit()

    print("✅ Email sent with ML report attached")


# =====================================================
# MAIN EXECUTION
# =====================================================
if __name__ == "__main__":

    # --- Existing pipeline ---
    remove_duplicates()
    config_filter()
    gp_merge()
    extract_date_time()
    fill_blanks()

    # --- Load cleaned data for ML ---
    print("\n📂 Loading cleaned data for ML processing...")
    df_clean = pd.read_excel(FINAL_FILE, dtype=str)

    # --- AI/ML pipeline ---
    df_features   = feature_engineering(df_clean)
    predictions   = predict_referral_trends(df_features)
    clusters      = cluster_doctors(df_features)
    anomalies     = detect_anomalies(df_features)

    # --- Save & notify ---
    save_ml_results(predictions, clusters, anomalies)

    add_graphs_to_excel(ML_FILE)
    create_dashboard(ML_FILE)   # 🔥 ADD THIS

    send_email()

    print("\n🎉 AI/ML PIPELINE COMPLETE 🎉")
    print(f"   → Cleaned data  : {FINAL_FILE}")
    print(f"   → ML output     : {ML_FILE}")
